import os
import csv
import time
import sys
import requests
import pandas as pd
import traceback
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ==========================================
# CONFIGURAÇÕES GLOBAIS
# ==========================================
CORES_STATUS = {
    "OK": "66ff66",                            # Verde
    "FAULT": "ff6666",                         # Vermelho
    "CONFIRMED": "ffff66",                     # Amarelo
    "OK (prev. fault unconfirmed)": "6666ff"  # Azul
}

MAPA_STATUS = {
    0: 'OK',
    1: 'FAULT',
    2: 'CONFIRMED',
    3: 'OK (prev. fault unconfirmed)'
}

# 1. Configuração Inicial de Pastas e Ficheiro de Log
pasta_atual = os.getcwd()
data_atual = datetime.now().strftime("%d-%m-%Y")
nome_pasta_logs = f'Logs_{data_atual}'

# Garante que a pasta do dia existe antes de abrir o log
if not os.path.exists(nome_pasta_logs):
    os.mkdir(nome_pasta_logs)

# Redireciona o stdout e stderr para o ficheiro log.txt dentro da pasta gerada
caminho_log = os.path.join(nome_pasta_logs, 'log.txt')
arquivo_log = open(caminho_log, 'a', encoding='utf-8')
sys.stdout = arquivo_log
sys.stderr = arquivo_log

print(f"\n========================================================")
print(f"=== INICIANDO EXECUÇÃO SILENCIOSA: {datetime.now().strftime('%H:%M:%S')} ===")
print(f"========================================================")
print(f"[DEBUG] Diretoria atual: {pasta_atual}")
print(f"[DEBUG] Pasta de destino dos logs: {nome_pasta_logs}")

# Tenta remover o ficheiro temporário antigo se ele existir
if os.path.exists('data.xlsx'):
    try:
        os.remove('data.xlsx')
        print("[DEBUG] Ficheiro temporário 'data.xlsx' antigo removido.")
    except Exception as e:
        print(f"[AVISO] Não foi possível remover 'data.xlsx': {e}")

# 2. Configuração do Edge Headless (Segundo Plano)
config_edge = Options()
config_edge.add_argument("--headless=new")
config_edge.add_argument("--disable-gpu")
config_edge.add_argument("--window-size=1920,1080")
config_edge.add_argument("--disable-infobars")
config_edge.add_argument("--no-sandbox")

try:
    # 3. Leitura das credenciais (acesso.txt)
    site, username, password = None, None, None
    print("[DEBUG] A ler o ficheiro 'acesso.txt'...")
    
    if not os.path.exists('acesso.txt'):
        print("[ERRO CRÍTICO] O ficheiro 'acesso.txt' não foi encontrado.")
        sys.exit(1)

    with open('acesso.txt', 'r', encoding='utf-8') as file:
        for line in file:
            parts = line.split("--")
            if len(parts) == 2:
                key = parts[0].strip()
                val = parts[1].strip()
                if key == "Site":
                    site = val
                elif key == "Username":
                    username = val
                elif key == "Password":
                    password = val
                    
    print(f"[DEBUG] Credenciais lidas -> Site: {site}, Username: {username}")

    if not all([site, username, password]):
        print("[ERRO CRÍTICO] 'acesso.txt' tem dados em falta.")
        sys.exit(1)

    # 4. Autenticação no Navegador
    print("\n[DEBUG] A iniciar o Edge em segundo plano...")
    try:
        servico_edge = Service()
        driver = webdriver.Edge(service=servico_edge, options=config_edge)
    except Exception as err_nav:
        print(f"[ERRO CRÍTICO] Falha ao iniciar o Edge:\n{traceback.format_exc()}")
        sys.exit(1)

    wait = WebDriverWait(driver, 15)

    url_visu = f"http://{site}/web.visu/"
    print(f"[DEBUG] A abrir URL de Login: {url_visu}")
    driver.get(url_visu)
    time.sleep(3)

    print("[DEBUG] A clicar em 'Manual Login'...")
    botao_manual = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'userButton') and contains(text(), 'Manual Login')]")))
    driver.execute_script("arguments[0].click();", botao_manual)
    time.sleep(4)

    print("[DEBUG] A preencher credenciais...")
    campo_user = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='text'] | //form//input[1]")))
    campo_user.clear()
    campo_user.send_keys(username)

    campo_pass = driver.find_element(By.XPATH, "//input[@type='password'] | //form//input[2]")
    campo_pass.clear()
    campo_pass.send_keys(password)

    print("[DEBUG] A submeter formulário...")
    try:
        botao_confirmar = driver.find_element(By.XPATH, "//form/div[4] | //div[contains(@class, 'button') and text()='OK']")
        driver.execute_script("arguments[0].click();", botao_confirmar)
    except Exception:
        print("[DEBUG] Botão OK não encontrado, a tentar enviar ENTER...")
        campo_pass.send_keys(Keys.ENTER)

    print("[DEBUG] A aguardar validação do loginId na URL...")
    try:
        WebDriverWait(driver, 25).until(EC.url_contains("loginId="))
    except Exception:
        print("[AVISO] Tempo limite de espera esgotado.")

    url_pos_login = driver.current_url
    print(f"[DEBUG] URL detectada após login: {url_pos_login}")
    
    if "loginId=" in url_pos_login:
        id_login = url_pos_login.split("loginId=")[1].split("&")[0]
        print(f"\n[SUCESSO] Token de Sessão obtido: {id_login}")
    else:
        print(f"\n[ERRO CRÍTICO] Não conseguiu extrair o loginId da URL.")
        driver.quit()
        sys.exit(1)

    driver.quit()
    print("[DEBUG] Navegador fechado. Iniciando fase API Requests.")

    # 5. Processamento dos UUIDs via API (Requests + Pandas)
    if not os.path.exists('uuids.csv'):
        print("[ERRO CRÍTICO] O ficheiro 'uuids.csv' não foi encontrado.")
        sys.exit(1)

    print("\n[DEBUG] A abrir o ficheiro 'uuids.csv'...")
    with open('uuids.csv', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile, delimiter=';')
        linhas = list(reader)
        print(f"[DEBUG] Total de linhas no CSV: {len(linhas)}")
        
        for num_linha, row in enumerate(linhas, start=1):
            if not row:
                continue
            
            if len(row) < 2:
                row = row[0].split(',')
                if len(row) < 2:
                    print(f"[ERRO] Linha {num_linha} inválida. Ignorada.")
                    continue

            uuid = row[0].strip()
            quarto = row[1].strip()
            
            url_api = f"http://{site}/webif/gaobjcgi?action=jobCommand&uuid={uuid}&cmd=getLog&loginId={id_login}"
            print(f"\n----------------------------------------")
            print(f"[DEBUG] A processar Linha {num_linha} -> Quarto: {quarto}")
            
            try:
                response = requests.get(url_api, timeout=15)
                if response.status_code == 200:
                    json_data = response.json()
                    
                    if 'history' in json_data and json_data['history']:
                        df = pd.json_normalize(json_data['history'])
                        
                        if 'timestamp' in df.columns:
                            df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms', errors='coerce').dt.tz_localize(None)
                        
                        if 'statusFrom' in df.columns:
                            df['statusFrom'] = df['statusFrom'].map(MAPA_STATUS).fillna(df['statusFrom'])
                        if 'statusTo' in df.columns:
                            df['statusTo'] = df['statusTo'].map(MAPA_STATUS).fillna(df['statusTo'])
                        
                        ficheiro_temporario = 'data.xlsx'
                        df.to_excel(ficheiro_temporario, index=False)
                        
                        nome_ficheiro_final = os.path.join(nome_pasta_logs, f'{quarto}_{data_atual}.xlsx')
                        
                        if os.path.exists(nome_ficheiro_final):
                            os.remove(nome_ficheiro_final)
                            
                        os.rename(ficheiro_temporario, nome_ficheiro_final)
                        
                        # Aplicar as cores nas células
                        wb = load_workbook(nome_ficheiro_final)
                        ws = wb.active
                        
                        for row_cells in ws.iter_rows(min_row=2):
                            for cell in row_cells:
                                if cell.value in CORES_STATUS:
                                    cor_hex = CORES_STATUS[cell.value]
                                    cell.fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
                        
                        wb.save(nome_ficheiro_final)
                        print(f"[SUCESSO] -> {quarto} gerado perfeitamente.")
                    else:
                        print(f"[AVISO] -> O histórico para {quarto} está vazio.")
                else:
                    print(f"[ERRO] Falha na API. Código HTTP: {response.status_code}")
                    
            except Exception as e_loop:
                print(f"[ERRO DE LOOP] Falha na linha do quarto {quarto}:")
                print(traceback.format_exc())

    print(f"\n========================================")
    print(f"Processo finalizado com sucesso! Pasta: {nome_pasta_logs}")
    print(f"========================================")

except Exception as e_geral:
    print(f"\n[ERRO GERAL DO SISTEMA]:")
    print(traceback.format_exc())

finally:
    print("\n--- FIM DA EXECUÇÃO ---")
    # Garante que o arquivo fecha de forma segura salvando todos os dados em disco
    arquivo_log.close()
